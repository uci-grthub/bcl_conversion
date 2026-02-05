"""
Metadata validation module for Excel-based sequencing metadata files.
This module provides validation, detection, and highlighting of issues in metadata.
"""
import os
import re
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except Exception:
    openpyxl = None


def validate_metadata_and_write_report(metadata_file, out_xlsx=None):
    """Validate metadata Excel and write a highlighted copy + change list.

    - out_xlsx: path to write validation workbook (xlsx). If openpyxl not available
      a plain text log will be written to logs/metadata_validation.txt instead.
    """
    issues = []
    seq_pattern = re.compile(r'^[ACGTNacgtn]+$')

    def _is_sequence_series(series):
        vals = [str(v).strip() for v in series if not pd.isna(v)]
        vals = [v for v in vals if v and v.lower() != 'nan' and v.lower() != 'none']
        if not vals:
            return False
        return all(seq_pattern.match(v) for v in vals)

    def _norm_key(val):
        if pd.isna(val):
            return None
        try:
            num = float(val)
            if num.is_integer():
                return str(int(num))
        except Exception:
            pass
        return str(val).strip()

    if not metadata_file or not os.path.exists(metadata_file):
        issues.append({'sheet': '', 'row': '', 'col': '', 'message': f'Metadata file not found: {metadata_file}'})
        os.makedirs('logs', exist_ok=True)
        with open('logs/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: {it['row']} {it['col']} - {it['message']}\n")
        return

    try:
        xlf = pd.ExcelFile(metadata_file)
    except Exception as e:
        issues.append({'sheet': '', 'row': '', 'col': '', 'message': f'Could not open Excel: {e}'})
        os.makedirs('logs', exist_ok=True)
        with open('logs/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: {it['row']} {it['col']} - {it['message']}\n")
        return

    sheet_dfs = {}
    header_rows = {}
    barcode_len_by_lane_group = {}

    for sheet in xlf.sheet_names:
        try:
            raw = pd.read_excel(metadata_file, sheet_name=sheet, header=None)
        except Exception:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Could not read sheet'})
            continue

        # Find header row BEFORE filling (to avoid filling header rows into data rows)
        header_row = -1
        nrows = raw.shape[0]
        header_keywords = ['Lane', 'Sample_Project', 'Project name', 'Sample Name', 'Sample_Name', 'Sample_ID', 'Lab ID', 'Order ID', 'Email', 'Group', 'group', 'Gr']
        for i in range(nrows):
            row = raw.iloc[i]
            row_vals = [str(x) if not pd.isna(x) else '' for x in row.values]

            # Direct match in single row
            if 'Lane' in row_vals and any(k in row_vals for k in ['Sample_Project', 'Project name', 'Sample Name', 'Sample_Name']):
                header_row = i
                break

            # Try combining with the next one or two rows to handle split/multi-line headers
            combined_vals = list(row_vals)
            for j in (1, 2):
                if i + j < nrows:
                    next_row = raw.iloc[i + j]
                    next_vals = [str(x) if not pd.isna(x) else '' for x in next_row.values]
                    combined_vals.extend(next_vals)
                    if 'Lane' in combined_vals and any(k in combined_vals for k in ['Sample_Project', 'Project name', 'Sample Name', 'Sample_Name', 'Lab ID', 'Order ID', 'Email', 'group', 'Group']):
                        # Prefer the lower row index if it contains the majority of header keywords
                        matches_curr = sum(1 for k in header_keywords if k in row_vals)
                        matches_next = sum(1 for k in header_keywords if k in next_vals)
                        header_row = i + j if matches_next >= matches_curr else i
                        break
                else:
                    break
            if header_row != -1:
                break

        if header_row == -1:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Header row not found'})
            # store raw as-is to copy into workbook
            df_sheet = raw.copy()
            sheet_dfs[sheet] = df_sheet
            header_rows[sheet] = None
            continue

        try:
            df = pd.read_excel(metadata_file, sheet_name=sheet, header=header_row)
        except Exception as e:
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': f'Error parsing sheet with detected header: {e}'})
            sheet_dfs[sheet] = raw
            header_rows[sheet] = None
            continue

        # After reading with proper header, fill down missing values in key columns
        # This handles merged cells in the Excel file where multiple rows share the same project/sample info
        try:
            # Identify columns that should be filled down (metadata columns, not barcode sequences)
            fill_cols = ['Lane', 'Group', 'group', 'Gr', 'Order ID', 'LabID', 'Lab ID', 'Contact', 'Email',
                        'Project name', 'Project', 'Sample_Project', 'Sample Name', 'Sample_Name',
                        'Sample_ID', 'Index Name']
            cols_to_fill = [c for c in fill_cols if c in df.columns]

            # Only forward-fill i7/index (first index) barcode columns when they contain sequence values.
            # Do NOT forward-fill i5/index2 (second index) columns to avoid propagating blanks for single-index libraries.
            seq_candidate_cols = ['index', 'i7 Barcode Sequence', 'Index']
            seq_cols = [c for c in seq_candidate_cols if c in df.columns and _is_sequence_series(df[c])]

            # Determine if this sheet is a Summary-like sheet; summary sheets should not
            # have Order ID forward-filled — instead, prefer per-row Lab ID values.
            is_summary_sheet = False
            try:
                if isinstance(sheet, str) and 'summary' in sheet.lower():
                    is_summary_sheet = True
            except Exception:
                is_summary_sheet = False

            # If summary sheet, exclude Order ID from forward-fill to avoid copying
            # one Order ID down the whole column; otherwise include it.
            if is_summary_sheet and 'Order ID' in cols_to_fill:
                cols_to_fill = [c for c in cols_to_fill if c != 'Order ID']

            final_fill = cols_to_fill + seq_cols
            if final_fill:
                df[final_fill] = df[final_fill].ffill()

            # After forward-fill: handle Order ID fills.
            try:
                lab_cols = []
                if 'Lab ID' in df.columns:
                    lab_cols.append('Lab ID')
                if 'LabID' in df.columns:
                    lab_cols.append('LabID')

                if 'Order ID' in df.columns and lab_cols:
                    if is_summary_sheet:
                        # For summary sheets, fill missing Order ID per-row from Lab ID values
                        for lc in lab_cols:
                            try:
                                # rows where Order ID is blank but lab id present
                                missing_order = df['Order ID'].isna() | (df['Order ID'].astype(str).str.strip() == '')
                                has_lab = ~(df[lc].isna()) & (df[lc].astype(str).str.strip() != '')
                                to_fill = missing_order & has_lab
                                if to_fill.any():
                                    df.loc[to_fill, 'Order ID'] = df.loc[to_fill, lc].astype(str).str.strip()
                                # do NOT break: allow other lab cols to supplement remaining blanks
                            except Exception:
                                continue
                    else:
                        # For non-summary sheets, backfill Order ID from Lab ID if entirely missing
                        for lc in lab_cols:
                            try:
                                missing_order = df['Order ID'].isna() | (df['Order ID'].astype(str).str.strip() == '')
                                has_lab = ~(df[lc].isna()) & (df[lc].astype(str).str.strip() != '')
                                to_fill = missing_order & has_lab
                                if to_fill.any():
                                    df.loc[to_fill, 'Order ID'] = df.loc[to_fill, lc].astype(str).str.strip()
                                    break
                            except Exception:
                                continue
            except Exception:
                pass
        except Exception:
            # If filling fails, continue with unfilled data
            pass

        # Make Sample_Name values unique within each project by appending suffixes
        # This handles cases where multiple barcodes belong to the same sample (merged cells)
        try:
            project_col = None
            if 'Project' in df.columns:
                project_col = 'Project'
            elif 'Project name' in df.columns:
                project_col = 'Project name'
            elif 'Sample_Project' in df.columns:
                project_col = 'Sample_Project'
            
            sample_name_col = None
            if 'Sample_Name' in df.columns:
                sample_name_col = 'Sample_Name'
            elif 'Sample Name' in df.columns:
                sample_name_col = 'Sample Name'
            
            # If we have both project and sample name columns, make sample names unique
            if project_col and sample_name_col:
                for project in df[project_col].unique():
                    if pd.isna(project) or str(project).strip() == '' or str(project).lower() == 'nan':
                        continue
                    
                    project_mask = df[project_col] == project
                    project_indices = df[project_mask].index
                    
                    # Count occurrences of each Sample_Name within this project
                    sample_names_in_project = df.loc[project_indices, sample_name_col]
                    sample_name_counts = sample_names_in_project.value_counts()
                    
                    # For Sample_Names that appear more than once, add suffixes
                    for sample_name, count in sample_name_counts.items():
                        if count > 1 and pd.notna(sample_name):
                            # Find all occurrences of this Sample_Name in this project
                            dup_mask = (df[project_col] == project) & (df[sample_name_col] == sample_name)
                            dup_indices = df[dup_mask].index
                            
                            # Append suffix to each duplicate (_1, _2, etc.)
                            for i, idx in enumerate(dup_indices, start=1):
                                df.loc[idx, sample_name_col] = f"{sample_name}_{i}"
        except Exception:
            # If uniqueness logic fails, continue with filled but non-unique sample names
            pass

        # Normalize project-name columns (handle minor differences like underscores, extra spaces, case)
        try:
            proj_variants = ['Project name', 'Project Name', 'Project', 'Sample_Project']
            present = [c for c in proj_variants if c in df.columns]
            if present:
                def _norm_proj_val(v):
                    if pd.isna(v):
                        return ''
                    s = str(v).strip()
                    s = s.replace('_', ' ')
                    s = ' '.join(s.split())
                    return s

                # create normalized versions and detect per-row inconsistencies
                norm_cols = {}
                for c in present:
                    norm_name = f"{c}__norm"
                    df[norm_name] = df[c].apply(_norm_proj_val)
                    norm_cols[c] = norm_name

                for ridx, row in df.iterrows():
                    vals = [row[nc] for nc in norm_cols.values() if row[nc] and str(row[nc]).strip() != '']
                    vals = list(dict.fromkeys(vals))
                    if len(vals) > 1:
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'Project', 'message': 'Inconsistent Project name across columns', 'lane': row.get('Lane', ''), 'group': row.get('Group', row.get('Gr', row.get('group', ''))), 'excel_row': int(ridx) + 2})

                # Create/overwrite a canonical 'Project' column with the first non-empty normalized value
                def _first_nonempty_norm(row):
                    for nc in norm_cols.values():
                        v = row.get(nc)
                        if v and str(v).strip() != '':
                            return v
                    return ''

                df['Project'] = df.apply(_first_nonempty_norm, axis=1)

                # drop temporary norm columns
                try:
                    df.drop(columns=list(norm_cols.values()), inplace=True)
                except Exception:
                    pass
        except Exception:
            pass

        # Build barcode length map by lane/group if barcode sequences are available
        try:
            lane_col = 'Lane' if 'Lane' in df.columns else None
            group_col = None
            if 'Group' in df.columns:
                group_col = 'Group'
            elif 'Gr' in df.columns:
                group_col = 'Gr'
            elif 'group' in df.columns:
                group_col = 'group'

            i7_col = None
            if 'index' in df.columns:
                i7_col = 'index'
            elif 'i7 Barcode Sequence' in df.columns:
                i7_col = 'i7 Barcode Sequence'
            elif 'Index' in df.columns and _is_sequence_series(df['Index']):
                i7_col = 'Index'

            i5_col = None
            if 'index2' in df.columns:
                i5_col = 'index2'
            elif 'i5 Barcode Sequence' in df.columns:
                i5_col = 'i5 Barcode Sequence'
            elif 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                i5_col = 'Index2'

            def _seq_len(series):
                vals = [str(v).strip() for v in series if not pd.isna(v)]
                vals = [v for v in vals if v and v.lower() != 'nan' and v.lower() != 'none']
                if not vals:
                    return 0
                lengths = {len(v) for v in vals}
                return max(lengths)

            if lane_col and group_col and (i7_col or i5_col):
                for (lane_val, group_val), sub in df.groupby([lane_col, group_col]):
                    lane_key = _norm_key(lane_val)
                    group_key = _norm_key(group_val)
                    if lane_key is None or group_key is None:
                        continue
                    i7_len = _seq_len(sub[i7_col]) if i7_col else 0
                    i5_len = _seq_len(sub[i5_col]) if i5_col else 0
                    barcode_len_by_lane_group[(lane_key, group_key)] = {
                        'i7_len': i7_len,
                        'i5_len': i5_len
                    }
        except Exception:
            pass

        sheet_dfs[sheet] = df
        header_rows[sheet] = header_row

        # Basic checks
        if 'Lane' not in df.columns:
            issues.append({'sheet': sheet, 'row': '', 'col': 'Lane', 'message': 'Missing Lane column'})
        else:
            # non-numeric lanes
            try:
                bad_lane = df[pd.to_numeric(df['Lane'], errors='coerce').isna()]
                if not bad_lane.empty:
                    for ridx in bad_lane.index.tolist():
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'Lane', 'message': 'Non-numeric Lane value'})
            except Exception:
                pass

        # Project column existence
        if not any(c in df.columns for c in ['Sample_Project', 'Project name', 'Project']):
            issues.append({'sheet': sheet, 'row': '', 'col': '', 'message': 'Missing Project column (Sample_Project or Project name)'})

        # Duplicate combined barcodes (only flag if on same lane)
        if 'index' in df.columns and 'Lane' in df.columns:
            idx1 = df['index'].fillna('').astype(str)
            idx2 = df['index2'].fillna('').astype(str) if 'index2' in df.columns else pd.Series([''] * len(df))
            combined = (idx1 + ':' + idx2).replace('nan', '')
            lanes = df['Lane'].fillna('').astype(str)
            
            # Find duplicate barcodes on same lane
            for lane in lanes.unique():
                if lane == '' or pd.isna(lane):
                    continue
                lane_mask = (lanes == lane)
                lane_combined = combined[lane_mask]
                dup_in_lane = lane_combined.duplicated(keep=False) & (lane_combined != ':')
                if dup_in_lane.any():
                    lane_dup_idxs = df.index[lane_mask][dup_in_lane].tolist()
                    for ridx in lane_dup_idxs:
                        issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'index', 'message': 'Duplicate barcode combination (index+index2) on same lane'})

        # Missing indexes when others exist
        if 'index' in df.columns:
            has_any = df['index'].notna() & (df['index'].astype(str).str.strip() != '')
            if has_any.any() and (~has_any).any():
                for ridx in df.index[~has_any].tolist():
                    issues.append({'sheet': sheet, 'row': int(ridx), 'col': 'index', 'message': 'Missing index while other rows have index'})

        # Masking validation runs after all sheets are processed

    # Validate masking against index lengths after all sheets are processed
    def _mask_len_map(masking_str):
        parts = re.split(r'[;,]', str(masking_str))
        mask_map = {}
        for part in parts:
            part = part.strip()
            if not part:
                continue
            m = re.match(r'^([A-Za-z0-9]+)\s*:\s*(\d+)$', part)
            if m:
                mask_map[m.group(1).upper()] = int(m.group(2))
        return mask_map

    def _get_barcode_len(lane_val, group_val):
        # Try multiple variants of lane/group keys to find a mapping in barcode_len_by_lane_group
        if lane_val is None or group_val is None:
            return None
        lk = _norm_key(lane_val)
        gk = _norm_key(group_val)
        candidates = []
        if lk is not None and gk is not None:
            candidates.append((lk, gk))
        try:
            candidates.append((str(lane_val).strip(), str(group_val).strip()))
        except Exception:
            pass
        try:
            candidates.append((str(float(lane_val)), str(float(group_val))))
        except Exception:
            pass
        for ck in candidates:
            if ck in barcode_len_by_lane_group:
                return barcode_len_by_lane_group.get(ck)
        return None

    for sheet, df in sheet_dfs.items():
        if 'Masking' not in df.columns:
            continue

        # iterate with positional index to avoid label/loc mismatches
        for pos, (ridx, row) in enumerate(df.iterrows()):
            masking_val = row.get('Masking')
            if pd.isna(masking_val) or str(masking_val).strip() == '':
                continue

            mask_map = _mask_len_map(masking_val)
            if not mask_map:
                issues.append({'sheet': sheet, 'row': int(pos), 'col': 'Masking', 'message': 'Unrecognized Masking format', 'lane': '', 'group': '', 'excel_row': int(pos) + 2})
                continue

            i1_len = mask_map.get('I1')
            i2_len = mask_map.get('I2')

            # Determine whether this Summary/DataFrame provides actual sequence values
            summary_has_seq1 = False
            summary_has_seq2 = False
            try:
                if 'index' in df.columns and _is_sequence_series(df['index']):
                    summary_has_seq1 = True
                if 'i7 Barcode Sequence' in df.columns and _is_sequence_series(df['i7 Barcode Sequence']):
                    summary_has_seq1 = True
                if 'Index' in df.columns and _is_sequence_series(df['Index']):
                    summary_has_seq1 = True

                if 'index2' in df.columns and _is_sequence_series(df['index2']):
                    summary_has_seq2 = True
                if 'i5 Barcode Sequence' in df.columns and _is_sequence_series(df['i5 Barcode Sequence']):
                    summary_has_seq2 = True
                if 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                    summary_has_seq2 = True
            except Exception:
                summary_has_seq1 = summary_has_seq1
                summary_has_seq2 = summary_has_seq2

            # Prefer barcode lengths from Barcode List via lane/group mapping
            lane_val = row.get('Lane') if 'Lane' in df.columns else None
            group_val = None
            if 'Group' in df.columns:
                group_val = row.get('Group')
            elif 'Gr' in df.columns:
                group_val = row.get('Gr')
            elif 'group' in df.columns:
                group_val = row.get('group')

            mapped = _get_barcode_len(lane_val, group_val)
            lane_key = None
            group_key = None
            if mapped is not None:
                # prefer normalized keys for reporting
                try:
                    lane_key = _norm_key(lane_val)
                    group_key = _norm_key(group_val)
                except Exception:
                    lane_key = lane_val
                    group_key = group_val
                len1 = mapped.get('i7_len', 0)
                len2 = mapped.get('i5_len', 0)
                # If mapped i5 length is zero but Barcode List sheet actually contains i5 sequences,
                # try to recompute lengths directly from the Barcode List sheet as a fail-safe.
                if len2 == 0:
                    try:
                        # find a sheet that looks like Barcode List
                        bdf = None
                        for sname, sdf in sheet_dfs.items():
                            if sname.lower().strip().startswith('barcode') or 'i7 Barcode Sequence' in sdf.columns or 'i5 Barcode Sequence' in sdf.columns:
                                bdf = sdf
                                break
                        if bdf is not None:
                            b_lane_col = 'Lane' if 'Lane' in bdf.columns else None
                            b_group_col = 'Group' if 'Group' in bdf.columns else ('Gr' if 'Gr' in bdf.columns else ('group' if 'group' in bdf.columns else None))
                            b_i5_col = 'i5 Barcode Sequence' if 'i5 Barcode Sequence' in bdf.columns else ('index2' if 'index2' in bdf.columns else ('Index2' if 'Index2' in bdf.columns else None))
                            if b_lane_col and b_group_col and b_i5_col:
                                try:
                                    lk = _norm_key(lane_val)
                                    gk = _norm_key(group_val)
                                    vals = []
                                    for _, brow in bdf.iterrows():
                                        try:
                                            if _norm_key(brow[b_lane_col]) == lk and _norm_key(brow[b_group_col]) == gk:
                                                v = brow[b_i5_col]
                                                if not pd.isna(v):
                                                    sv = str(v).strip()
                                                    if sv and sv.lower() not in ('nan','none'):
                                                        vals.append(sv)
                                        except Exception:
                                            continue
                                    if vals:
                                        len2 = max(len(v) for v in vals)
                                except Exception:
                                    pass
                    except Exception:
                        pass
            # Debug: record mapping lookup details for troubleshooting
            try:
                with open('logs/masking_lookup_debug.txt', 'a') as dbg:
                    dbg.write(f"sheet={sheet} pos={pos} lane_val={lane_val} group_val={group_val} mapped={mapped}\n")
            except Exception:
                pass

            # If no mapping exists, compute index lengths from the Summary row values
            if mapped is None:
                index1 = ''
                index2 = ''
                if 'index' in df.columns:
                    index1 = '' if pd.isna(row.get('index')) else str(row.get('index')).strip()
                elif 'i7 Barcode Sequence' in df.columns:
                    index1 = '' if pd.isna(row.get('i7 Barcode Sequence')) else str(row.get('i7 Barcode Sequence')).strip()
                elif 'Index' in df.columns and _is_sequence_series(df['Index']):
                    index1 = '' if pd.isna(row.get('Index')) else str(row.get('Index')).strip()

                if 'index2' in df.columns:
                    index2 = '' if pd.isna(row.get('index2')) else str(row.get('index2')).strip()
                elif 'i5 Barcode Sequence' in df.columns:
                    index2 = '' if pd.isna(row.get('i5 Barcode Sequence')) else str(row.get('i5 Barcode Sequence')).strip()
                elif 'Index2' in df.columns and _is_sequence_series(df['Index2']):
                    index2 = '' if pd.isna(row.get('Index2')) else str(row.get('Index2')).strip()

                len1 = 0 if index1 in ('', 'nan', 'None') else len(index1)
                len2 = 0 if index2 in ('', 'nan', 'None') else len(index2)

            if i1_len is not None and len1 != i1_len:
                issues.append({
                    'sheet': sheet,
                    'row': int(pos),
                    'col': 'Masking',
                    'message': f"Masking I1:{i1_len} does not match index length {len1}",
                    'lane': lane_key if lane_key is not None else lane_val,
                    'group': group_key if group_key is not None else group_val,
                    'excel_row': int(pos) + 2
                })

            if i2_len is not None:
                if i2_len == 0 and len2 != 0:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:0 but index2 length is {len2}",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })
                elif i2_len > 0 and len2 != i2_len:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:{i2_len} does not match index2 length {len2}",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })
                elif mapped is None and i2_len > 0 and 'index2' not in df.columns and 'i5 Barcode Sequence' not in df.columns:
                    issues.append({
                        'sheet': sheet,
                        'row': int(pos),
                        'col': 'Masking',
                        'message': f"Masking I2:{i2_len} but index2 column is missing",
                        'lane': lane_key if lane_key is not None else lane_val,
                        'group': group_key if group_key is not None else group_val,
                        'excel_row': int(pos) + 2
                    })

    # Write validation workbook if possible
    os.makedirs('logs', exist_ok=True)
    if out_xlsx is None:
        out_xlsx = os.path.join('logs', f"metadata_validation_{os.path.basename(metadata_file)}.xlsx")

    if openpyxl is None:
        # fallback: dump issues to text file
        with open('logs/metadata_validation.txt', 'w') as tf:
            for it in issues:
                tf.write(f"{it['sheet']}: row={it['row']} col={it['col']} - {it['message']}\n")
        return

    # Use pandas to write sheets, then highlight rows with issues
    try:
        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            for sheet, df in sheet_dfs.items():
                # If df has numeric header rows (raw), write as-is
                try:
                    df.to_excel(writer, sheet_name=sheet, index=False)
                except Exception:
                    # fallback to writing raw values
                    pd.DataFrame(df).to_excel(writer, sheet_name=sheet, index=False)

            # Write recommended changes sheet
            # Enrich issues with lane/group where possible for clearer RECOMMENDED_CHANGES
            for it in issues:
                sh = it.get('sheet')
                r = it.get('row')
                # Prefer existing values if already provided (e.g., masking checks set these)
                it_lane = it.get('lane', '')
                it_group = it.get('group', '')
                # If missing, try to look up from the sheet by row index
                try:
                    if (not it_lane or pd.isna(it_lane)) and sh in sheet_dfs and isinstance(r, (int, float)):
                        df_lookup = sheet_dfs[sh]
                        if 'Lane' in df_lookup.columns and int(r) in df_lookup.index:
                            it_lane = df_lookup.loc[int(r), 'Lane']
                    if (not it_group or pd.isna(it_group)) and sh in sheet_dfs and isinstance(r, (int, float)):
                        df_lookup = sheet_dfs[sh]
                        for gcol in ('Group', 'Gr', 'group'):
                            if gcol in df_lookup.columns and int(r) in df_lookup.index:
                                it_group = df_lookup.loc[int(r), gcol]
                                break
                except Exception:
                    pass
                it['lane'] = it_lane if it_lane is not None else ''
                it['group'] = it_group if it_group is not None else ''
                # Prefer existing excel_row if present, otherwise compute
                try:
                    if 'excel_row' in it and it.get('excel_row') not in (None, ''):
                        pass
                    else:
                        it['excel_row'] = int(r) + 2 if isinstance(r, (int, float)) else ''
                except Exception:
                    it['excel_row'] = ''

            if issues:
                issues_df = pd.DataFrame(issues)
            else:
                issues_df = pd.DataFrame([{'sheet': 'OK', 'row': '', 'col': '', 'message': 'No issues detected'}])
            issues_df.to_excel(writer, sheet_name='RECOMMENDED_CHANGES', index=False)

        # Apply highlighting
        wb = openpyxl.load_workbook(out_xlsx)
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        bold = Font(bold=True)

        # Group issues by sheet and highlight entire row where issue occurred
        for it in issues:
            sh = it.get('sheet')
            r = it.get('row')
            if sh in wb.sheetnames and isinstance(r, (int, float)):
                ws = wb[sh]
                # Excel output from pandas has header at row 1 and data starting at row 2.
                # Map DataFrame index -> excel row by adding 2 (1-based excel rows).
                try:
                    excel_row = int(r) + 2
                except Exception:
                    continue
                max_col = ws.max_column
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=excel_row, column=c)
                    try:
                        cell.fill = red_fill
                    except Exception:
                        pass

        # Bold header rows where header detected
        for sh, hr in header_rows.items():
            if sh not in wb.sheetnames:
                continue
            ws = wb[sh]
            # Header written by pandas is at row 1 in the output workbook
            excel_header = 1
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=excel_header, column=c)
                    cell.font = bold
                except Exception:
                    pass

        wb.save(out_xlsx)
    except Exception as e:
        # final fallback: write issues to text file
        with open('logs/metadata_validation.txt', 'w') as tf:
            tf.write(f'Error writing validation workbook: {e}\n')
            for it in issues:
                tf.write(f"{it['sheet']}: row={it['row']} col={it['col']} - {it['message']}\n")
    return
